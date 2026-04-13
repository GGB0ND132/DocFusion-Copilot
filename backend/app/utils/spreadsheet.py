from __future__ import annotations

import io
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree as ET

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
WORKBOOK_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

NS = {
    "main": MAIN_NS,
    "rel": PACKAGE_REL_NS,
}

ET.register_namespace("", MAIN_NS)
ET.register_namespace("r", WORKBOOK_REL_NS)
ET.register_namespace("mc", "http://schemas.openxmlformats.org/markup-compatibility/2006")


def _register_namespaces_from_xml(raw_xml: bytes) -> None:
    """从原始 XML 中提取并注册所有命名空间前缀，确保重新序列化时保持一致。
    Extract and register all namespace prefixes from raw XML so re-serialization preserves them.
    """
    for _event, elem in ET.iterparse(io.BytesIO(raw_xml), events=["start-ns"]):
        prefix, uri = elem
        if prefix:
            try:
                ET.register_namespace(prefix, uri)
            except ValueError:
                pass

_CELL_REF_RE = re.compile(r"(?P<column>[A-Z]+)(?P<row>\d+)")


@dataclass(slots=True)
class SpreadsheetRow:
    """工作表中单行的内存表示。
    In-memory representation of one worksheet row.
    """

    row_index: int
    values: list[str]


@dataclass(slots=True)
class SpreadsheetSheet:
    """单个工作表的内存表示。
    In-memory representation of one worksheet.
    """

    name: str
    rows: list[SpreadsheetRow]


@dataclass(slots=True)
class SpreadsheetDocument:
    """XLSX 工作簿的内存表示。
    In-memory representation of an XLSX workbook.
    """

    sheets: list[SpreadsheetSheet]


@dataclass(slots=True)
class CellWrite:
    """单个工作簿单元格的计划写入操作。
    Planned write operation for one workbook cell.
    """

    sheet_name: str
    cell_ref: str
    value: str | float | int


def _q(tag: str) -> str:
    """为 SpreadsheetML 标签补齐主工作表命名空间。
    Qualify a SpreadsheetML tag with the main worksheet namespace.
    """
    return f"{{{MAIN_NS}}}{tag}"


def column_letters_to_index(column_letters: str) -> int:
    """将 Excel 列字母如 `AB` 转换为从 1 开始的列索引。
    Convert Excel column letters like `AB` into a 1-based index.
    """
    value = 0
    for char in column_letters:
        value = value * 26 + (ord(char) - ord("A") + 1)
    return value


def index_to_column_letters(index: int) -> str:
    """将从 1 开始的列索引转换为 Excel 列字母。
    Convert a 1-based column index into Excel letters.
    """
    result = []
    current = index
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        result.append(chr(ord("A") + remainder))
    return "".join(reversed(result))


def split_cell_ref(cell_ref: str) -> tuple[int, int]:
    """将单元格引用拆分为行号和列号。
    Split a cell reference into row and column indices.
    """
    match = _CELL_REF_RE.fullmatch(cell_ref.upper())
    if not match:
        raise ValueError(f"Invalid cell reference: {cell_ref}")
    column_index = column_letters_to_index(match.group("column"))
    row_index = int(match.group("row"))
    return row_index, column_index


def build_cell_ref(row_index: int, column_index: int) -> str:
    """根据行列索引构造 Excel 单元格引用。
    Build an Excel cell reference from row and column indices.
    """
    return f"{index_to_column_letters(column_index)}{row_index}"


def load_xlsx(path: str | Path) -> SpreadsheetDocument:
    """读取精简版 XLSX 工作簿到内存工作表和行对象中。
    Read a minimal XLSX workbook into in-memory sheet and row objects.
    """
    workbook_path = Path(path)
    with zipfile.ZipFile(workbook_path, "r") as archive:
        shared_strings = _load_shared_strings(archive)
        sheet_targets = _load_sheet_targets(archive)
        sheets: list[SpreadsheetSheet] = []

        for sheet_name, target in sheet_targets:
            root = ET.fromstring(archive.read(target))
            sheet_data = root.find("main:sheetData", NS)
            rows: list[SpreadsheetRow] = []
            if sheet_data is None:
                sheets.append(SpreadsheetSheet(name=sheet_name, rows=[]))
                continue

            row_maps: list[tuple[int, dict[int, str]]] = []
            for row_el in sheet_data.findall("main:row", NS):
                row_index = int(row_el.get("r", "0"))
                value_map: dict[int, str] = {}
                for cell_el in row_el.findall("main:c", NS):
                    cell_ref = cell_el.get("r")
                    if not cell_ref:
                        continue
                    _, column_index = split_cell_ref(cell_ref)
                    value_map[column_index] = _read_cell_value(cell_el, shared_strings)
                row_maps.append((row_index, value_map))

            for row_index, value_map in row_maps:
                row_max_column = max(value_map.keys(), default=0)
                values = [value_map.get(column_index, "") for column_index in range(1, row_max_column + 1)]
                rows.append(SpreadsheetRow(row_index=row_index, values=values))

            sheets.append(SpreadsheetSheet(name=sheet_name, rows=rows))

    return SpreadsheetDocument(sheets=sheets)


def apply_xlsx_updates(
    template_path: str | Path,
    output_path: str | Path,
    updates: list[CellWrite],
) -> None:
    """使用 openpyxl 应用工作簿单元格更新，确保输出文件完全符合 OOXML 规范。
    Apply cell updates to a workbook via openpyxl, producing a spec-compliant output.
    """
    from openpyxl import load_workbook as _load_wb

    template_file = Path(template_path)
    destination_file = Path(output_path)
    destination_file.parent.mkdir(parents=True, exist_ok=True)

    wb = _load_wb(template_file)

    for update in updates:
        if update.sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{update.sheet_name}' not found in workbook.")
        ws = wb[update.sheet_name]
        cell = ws[update.cell_ref.upper()]
        if isinstance(update.value, (int, float)) and not isinstance(update.value, bool):
            cell.value = update.value
        else:
            cell.value = str(update.value)

    wb.save(destination_file)


def _load_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    """如果存在，则从工作簿压缩包加载共享字符串表。
    Load the shared-string table from a workbook archive if present.
    """
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    values: list[str] = []
    for item in root.findall("main:si", NS):
        text = "".join(node.text or "" for node in item.findall(".//main:t", NS))
        values.append(text)
    return values


def _load_sheet_targets(archive: zipfile.ZipFile) -> list[tuple[str, str]]:
    """解析工作簿中的工作表名称与 XML 路径映射。
    Resolve workbook sheet names to worksheet XML paths.
    """
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    relations = {}
    for relation in rels_root.findall("rel:Relationship", NS):
        target_path = relation.get("Target", "")
        if not target_path.startswith("xl/"):
            target_path = f"xl/{target_path}"
        relations[relation.get("Id")] = target_path

    targets: list[tuple[str, str]] = []
    for sheet_el in workbook_root.findall("main:sheets/main:sheet", {"main": MAIN_NS}):
        relation_id = sheet_el.get(f"{{{WORKBOOK_REL_NS}}}id")
        if not relation_id or relation_id not in relations:
            continue
        targets.append((sheet_el.get("name", "Sheet1"), relations[relation_id]))
    return targets


def _read_cell_value(cell_el: ET.Element, shared_strings: list[str]) -> str:
    """从工作表 XML 单元格元素中读取文本值。
    Read a textual cell value from a worksheet XML cell element.
    """
    cell_type = cell_el.get("t")
    value_node = cell_el.find("main:v", NS)
    if cell_type == "s" and value_node is not None and value_node.text is not None:
        index = int(value_node.text)
        return shared_strings[index] if index < len(shared_strings) else ""
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell_el.findall(".//main:t", NS))
    return value_node.text if value_node is not None and value_node.text is not None else ""


def _get_or_create_row(sheet_data: ET.Element, row_index: int) -> ET.Element:
    """返回指定索引对应的行元素，必要时创建并排序插入。
    Return the row element for an index, creating and ordering it if needed.
    """
    for row_el in sheet_data.findall("main:row", NS):
        if int(row_el.get("r", "0")) == row_index:
            return row_el

    new_row = ET.Element(_q("row"), {"r": str(row_index)})
    inserted = False
    for position, row_el in enumerate(sheet_data.findall("main:row", NS)):
        if int(row_el.get("r", "0")) > row_index:
            sheet_data.insert(position, new_row)
            inserted = True
            break
    if not inserted:
        sheet_data.append(new_row)
    return new_row


def _get_or_create_cell(row_el: ET.Element, cell_ref: str) -> ET.Element:
    """返回指定引用的单元格元素，不存在时则创建。
    Return the cell element for a reference, creating it if missing.
    """
    _, target_column = split_cell_ref(cell_ref)
    for cell_el in row_el.findall("main:c", NS):
        existing_ref = cell_el.get("r")
        if existing_ref == cell_ref:
            return cell_el

    new_cell = ET.Element(_q("c"), {"r": cell_ref})
    inserted = False
    for position, cell_el in enumerate(row_el.findall("main:c", NS)):
        existing_ref = cell_el.get("r")
        if existing_ref:
            _, existing_column = split_cell_ref(existing_ref)
            if existing_column > target_column:
                row_el.insert(position, new_cell)
                inserted = True
                break
    if not inserted:
        row_el.append(new_cell)
    return new_cell


def _set_cell_value(cell_el: ET.Element, value: str | float | int) -> None:
    """将 Python 值写入工作表单元格 XML 元素。
    Write a Python value into a worksheet cell element.
    """
    cell_ref = cell_el.get("r", "")
    style_id = cell_el.get("s")
    cell_el.attrib.clear()
    cell_el.set("r", cell_ref)
    if style_id is not None:
        cell_el.set("s", style_id)

    for child in list(cell_el):
        cell_el.remove(child)

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        value_node = ET.SubElement(cell_el, _q("v"))
        if isinstance(value, float) and value.is_integer():
            value_node.text = str(int(value))
        elif isinstance(value, float):
            value_node.text = f"{value:.6f}".rstrip("0").rstrip(".")
        else:
            value_node.text = str(value)
        return

    cell_el.set("t", "inlineStr")
    inline_str = ET.SubElement(cell_el, _q("is"))
    text_node = ET.SubElement(inline_str, _q("t"))
    text_node.text = str(value)


MC_NS = "http://schemas.openxmlformats.org/markup-compatibility/2006"

# Well-known OOXML namespace URIs that mc:Ignorable may reference.
_OOXML_KNOWN_NS: dict[str, str] = {
    "x14ac": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac",
    "xr": "http://schemas.microsoft.com/office/spreadsheetml/2014/revision",
    "xr2": "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2",
    "xr3": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3",
    "xr6": "http://schemas.microsoft.com/office/spreadsheetml/2014/revision6",
    "xr10": "http://schemas.microsoft.com/office/spreadsheetml/2014/revision10",
    "x16r2": "http://schemas.microsoft.com/office/spreadsheetml/2015/02/main",
}


def _patch_mc_ignorable(root: ET.Element, raw_xml: bytes) -> None:
    """确保 mc:Ignorable 中列出的所有命名空间前缀都有对应的声明。
    Ensure every prefix listed in mc:Ignorable has a matching xmlns declaration.

    Python's ElementTree strips unused namespace declarations during
    serialization. OOXML requires them to be present if referenced by
    mc:Ignorable, otherwise Excel reports content problems.
    """
    ignorable = root.get(f"{{{MC_NS}}}Ignorable")
    if not ignorable:
        return

    # Collect namespace URIs declared in the original XML
    original_ns: dict[str, str] = {}
    for _event, elem in ET.iterparse(io.BytesIO(raw_xml), events=["start-ns"]):
        prefix, uri = elem
        if prefix:
            original_ns[prefix] = uri

    # For each prefix in mc:Ignorable, ensure the root element carries its xmlns
    for prefix in ignorable.split():
        uri = original_ns.get(prefix) or _OOXML_KNOWN_NS.get(prefix)
        if uri:
            root.set(f"xmlns:{prefix}", uri)


def _update_dimension(root: ET.Element, sheet_data: ET.Element) -> None:
    """根据实际数据范围更新 worksheet 的 dimension 元素。
    Update the worksheet dimension element to reflect the actual data range.
    """
    max_row = 0
    max_col = 0
    for row_el in sheet_data.findall("main:row", NS):
        row_idx = int(row_el.get("r", "0"))
        if row_idx > max_row:
            max_row = row_idx
        for cell_el in row_el.findall("main:c", NS):
            ref = cell_el.get("r")
            if ref:
                _, col_idx = split_cell_ref(ref)
                if col_idx > max_col:
                    max_col = col_idx

    if max_row == 0 or max_col == 0:
        return

    new_ref = f"A1:{index_to_column_letters(max_col)}{max_row}"
    dim_el = root.find("main:dimension", NS)
    if dim_el is not None:
        dim_el.set("ref", new_ref)
    else:
        dim_el = ET.Element(_q("dimension"), {"ref": new_ref})
        root.insert(0, dim_el)
